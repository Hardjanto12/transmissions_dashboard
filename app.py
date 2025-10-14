from flask import Flask, render_template, request, jsonify, send_file
import ast
import copy
import logging
import os
import re
import json
import sys
from datetime import datetime
import glob
import socket
import threading
from logging.handlers import RotatingFileHandler
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import tempfile
from html import unescape
from pathlib import Path
import requests

BASE_DIR = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
TEMPLATE_FOLDER = BASE_DIR / "templates"
STATIC_FOLDER = BASE_DIR / "assets"

app = Flask(__name__, template_folder=str(TEMPLATE_FOLDER), static_folder=str(STATIC_FOLDER), static_url_path='/assets')
logger = logging.getLogger(__name__)
SHUTDOWN_TOKEN = os.environ.get('TRANSMISSION_SHUTDOWN_TOKEN', 'transmission-shutdown')

# Settings configuration
SETTINGS_FILE = 'settings.json'

FTP_TARGET_SLOTS = 2
DEFAULT_FTP_PORT = 21
DEFAULT_FTP_PING_INTERVAL = 60
DEFAULT_RESEND_TIMEOUT = 15
MAX_REMOTE_RESPONSE_PREVIEW = 1000

PING_LOG_MAX_BYTES = 10_000 * 1024 * 1024  # 10,000 MB limit for ping logs
PING_LOG_FILENAME = 'ping_status.log'

ping_logger = logging.getLogger('transmission.ping_status')
ping_logger.setLevel(logging.INFO)
ping_logger.propagate = False

def configure_ping_logger(logs_dir):
    """Configure rotating file handler for FTP ping status logs."""
    try:
        absolute_dir = os.path.abspath(logs_dir)
        os.makedirs(absolute_dir, exist_ok=True)
    except OSError as exc:  # pragma: no cover - filesystem dependent
        logger.exception('Failed to prepare ping log directory %s: %s', logs_dir, exc)
        return

    log_path = os.path.join(absolute_dir, PING_LOG_FILENAME)

    try:
        handler = RotatingFileHandler(
            log_path,
            maxBytes=PING_LOG_MAX_BYTES,
            backupCount=5,
            encoding='utf-8'
        )
    except Exception as exc:  # pragma: no cover - filesystem dependent
        logger.exception('Failed to initialize ping log handler for %s: %s', log_path, exc)
        return

    handler.setFormatter(logging.Formatter('%(asctime)s %(message)s'))

    for existing in list(ping_logger.handlers):
        ping_logger.removeHandler(existing)
        existing.close()

    ping_logger.addHandler(handler)

def sanitize_ftp_targets(targets, strict=False):
    """Normalize FTP targets to the expected structure."""
    normalized = []

    if not isinstance(targets, list):
        targets = []

    for index in range(FTP_TARGET_SLOTS):
        target = targets[index] if index < len(targets) else {}
        if not isinstance(target, dict):
            target = {}

        host = str(target.get('host', '') or '').strip()
        port_value = target.get('port', DEFAULT_FTP_PORT)

        if port_value in (None, ''):
            port_value = DEFAULT_FTP_PORT

        try:
            port = int(port_value)
        except (TypeError, ValueError):
            if strict:
                raise ValueError(f'FTP target {index + 1} port must be a number')
            port = DEFAULT_FTP_PORT

        if strict and not (1 <= port <= 65535):
            raise ValueError(f'FTP target {index + 1} port must be between 1 and 65535')

        if not strict and not (1 <= port <= 65535):
            port = DEFAULT_FTP_PORT

        normalized.append({
            'host': host,
            'port': port
        })

    return normalized


def sanitize_ping_interval(value, default_value):
    """Convert ping interval to an integer, falling back to default on error."""
    try:
        interval = int(value)
        if interval <= 0:
            raise ValueError
        return interval
    except (TypeError, ValueError):
        return default_value


def validate_ping_interval(value):
    """Validate and return a positive ping interval."""
    try:
        interval = int(value)
    except (TypeError, ValueError):
        raise ValueError('FTP ping interval must be a positive integer')

    if interval <= 0:
        raise ValueError('FTP ping interval must be a positive integer')

    return interval


CONTAINER_SEPARATOR_TOKEN_PATTERN = re.compile(r'[+\\]+')
CONTAINER_FIELD_TEXT_PATTERN = re.compile(
    r'(?P<prefix>(?:"|\')\s*container(?:_?no)?\s*(?:"|\')\s*:\s*(?:"|\'))'
    r'(?P<value>[^"\']*)'
    r'(?P<suffix>(?:"|\'))',
    re.IGNORECASE
)
CONTAINER_XML_FIELD_PATTERN = re.compile(
    r'(?P<prefix><\s*container(?:_?no)?\s*>)(?P<value>[^<]*)(?P<suffix></\s*container(?:_?no)?\s*>)',
    re.IGNORECASE
)


def normalize_container_separator_value(value):
    """Convert container separators to escaped backslashes suitable for resend."""
    if not isinstance(value, str):
        return value

    stripped = value.strip()
    if not stripped:
        return value

    if '+' not in stripped and '\\' not in stripped:
        return stripped

    segments = [segment for segment in CONTAINER_SEPARATOR_TOKEN_PATTERN.split(stripped) if segment]
    if not segments:
        return '\\' if stripped else stripped

    normalized = '\\'.join(segments)
    if stripped[-1] in ('+', '\\') and not normalized.endswith('\\'):
        normalized += '\\'
    return normalized


def normalize_containers_in_payload(payload):
    """Recursively normalize container separators within payload structures."""
    if isinstance(payload, dict):
        for key, value in payload.items():
            lowered_key = str(key).lower()
            if isinstance(value, (dict, list)):
                normalize_containers_in_payload(value)
                continue

            if isinstance(value, str):
                if 'container' in lowered_key and 'no' in lowered_key:
                    cleaned = normalize_container_separator_value(value)
                else:
                    cleaned = normalize_container_fields_in_text(value)
                if cleaned != value:
                    payload[key] = cleaned
        return payload

    if isinstance(payload, list):
        for index, item in enumerate(payload):
            if isinstance(item, (dict, list)):
                normalize_containers_in_payload(item)
                continue
            if isinstance(item, str):
                cleaned = normalize_container_fields_in_text(item)
                if cleaned != item:
                    payload[index] = cleaned

    return payload


def normalize_container_fields_in_text(payload_text, *, escape_for_literal=False):
    """Replace container separators inside raw payload text representations."""
    if not isinstance(payload_text, str):
        return payload_text

    def _replace_json(match):
        cleaned_value = normalize_container_separator_value(match.group('value'))
        if escape_for_literal:
            encoded_value = cleaned_value.replace('\\', '\\\\')
        else:
            encoded_value = cleaned_value
        return f"{match.group('prefix')}{encoded_value}{match.group('suffix')}"

    def _replace_xml(match):
        cleaned_value = normalize_container_separator_value(match.group('value'))
        if escape_for_literal:
            encoded_value = cleaned_value.replace('\\', '\\\\')
        else:
            encoded_value = cleaned_value
        return f"{match.group('prefix')}{encoded_value}{match.group('suffix')}"

    updated_text = CONTAINER_FIELD_TEXT_PATTERN.sub(_replace_json, payload_text)
    updated_text = CONTAINER_XML_FIELD_PATTERN.sub(_replace_xml, updated_text)
    return updated_text


def build_resend_url(server, endpoint):
    """Construct the resend URL from configured server and endpoint values."""
    server_value = (server or '').strip()
    endpoint_value = (endpoint or '').strip()

    if endpoint_value:
        endpoint_lower = endpoint_value.lower()
        if endpoint_lower.startswith(('http://', 'https://')):
            return endpoint_value

    if not server_value:
        raise ValueError('Resend server is not configured')

    normalized_server = server_value.rstrip('/')
    if not endpoint_value:
        return normalized_server

    normalized_endpoint = endpoint_value.lstrip('/')
    return f"{normalized_server}/{normalized_endpoint}"


def interpret_resend_response_success(response_obj, raw_text):
    """Determine whether a resend response represents a business success."""
    if response_obj is not None and not getattr(response_obj, 'ok', False):
        return False

    if raw_text is None:
        return False

    text = str(raw_text).strip()
    if not text:
        return False

    success_tokens = {'true', '1', 'yes', 'ok', 'success', 'berhasil', '200', 'completed', 'done'}
    failure_tokens = {'false', '0', 'no', 'failed', 'fail', 'error', 'gagal', 'not ok', 'failure'}
    success_substrings = {'success', 'berhasil', 'completed', 'done', 'ok'}
    failure_substrings = {'fail', 'failed', 'error', 'gagal', 'unsuccess', 'not ok', 'not success'}

    def classify_value(value):
        if isinstance(value, bool):
            return 'success' if value else 'failure'
        if isinstance(value, (int, float)):
            return 'success' if value else 'failure'
        if isinstance(value, str):
            lowered = value.strip().lower()
            if lowered in success_tokens:
                return 'success'
            if lowered in failure_tokens:
                return 'failure'
            if any(token in lowered for token in failure_substrings):
                return 'failure'
            if any(token in lowered for token in success_substrings) and not any(
                token in lowered for token in failure_substrings
            ):
                return 'success'
        return 'unknown'

    def evaluate_json(obj):
        success_found = False
        failure_found = False
        stack = [obj]

        while stack:
            current = stack.pop()
            if isinstance(current, dict):
                for key, value in current.items():
                    key_lower = str(key).lower()

                    if key_lower in ('resultcode', 'success', 'is_success', 'issuccess', 'successflag'):
                        classification = classify_value(value)
                    elif key_lower in ('status', 'result', 'response', 'state', 'message'):
                        classification = classify_value(value)
                    else:
                        classification = 'unknown'

                    if classification == 'success':
                        success_found = True
                    elif classification == 'failure':
                        failure_found = True

                    if isinstance(value, (dict, list)):
                        stack.append(value)
            elif isinstance(current, list):
                stack.extend(current)

        if failure_found and not success_found:
            return False
        if success_found and not failure_found:
            return True
        if success_found and failure_found:
            return False
        return False

    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        normalized = text.lower()
        if normalized in success_tokens:
            return True
        if any(token in normalized for token in failure_tokens.union(failure_substrings)):
            return False
        if any(token in normalized for token in success_substrings) and not any(
            token in normalized for token in failure_substrings
        ):
            return True
        return False

    return evaluate_json(parsed)


def build_initial_ftp_status_cache(targets):
    """Create an initial FTP status cache from configured targets."""
    statuses = []
    normalized = sanitize_ftp_targets(targets)

    for index, target in enumerate(normalized, start=1):
        status = 'unconfigured'
        if target['host']:
            status = 'unknown'

        statuses.append({
            'name': f'FTP Server {index}',
            'host': target['host'],
            'port': target['port'],
            'status': status,
            'error': None,
            'last_checked': None
        })

    return statuses


def load_settings():
    """Load settings from JSON file"""
    default_settings = {
        'logs_directory': 'logs',
        'auto_refresh_interval': 30,
        'ftp_targets': [
            {'host': '', 'port': DEFAULT_FTP_PORT},
            {'host': '', 'port': DEFAULT_FTP_PORT}
        ],
        'ftp_ping_interval': DEFAULT_FTP_PING_INTERVAL,
        'resend_server': '',
        'resend_endpoint': ''
    }

    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                settings = json.load(f)
                # Merge with defaults to ensure all keys exist
                for key, value in default_settings.items():
                    if key not in settings:
                        settings[key] = copy.deepcopy(value)

                settings['ftp_targets'] = sanitize_ftp_targets(
                    settings.get('ftp_targets', default_settings['ftp_targets'])
                )
                settings['ftp_ping_interval'] = sanitize_ping_interval(
                    settings.get('ftp_ping_interval'),
                    default_settings['ftp_ping_interval']
                )
                settings['resend_server'] = str(
                    settings.get('resend_server', '') or ''
                ).strip()
                settings['resend_endpoint'] = str(
                    settings.get('resend_endpoint', '') or ''
                ).strip()
                return settings
        except (json.JSONDecodeError, IOError):
            pass

    # Create default settings file
    save_settings(default_settings)
    return copy.deepcopy(default_settings)

def save_settings(settings):
    """Save settings to JSON file"""
    try:
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(settings, f, indent=2, ensure_ascii=False)
        return True
    except (IOError, TypeError):
        return False


ftp_status_lock = threading.Lock()
ftp_status_cache = []


# Load initial settings
app_settings = load_settings()
app_settings['ftp_targets'] = sanitize_ftp_targets(
    app_settings.get('ftp_targets'))
app_settings['ftp_ping_interval'] = sanitize_ping_interval(
    app_settings.get('ftp_ping_interval'), DEFAULT_FTP_PING_INTERVAL)
app_settings['resend_server'] = str(
    app_settings.get('resend_server', '') or ''
).strip()
app_settings['resend_endpoint'] = str(
    app_settings.get('resend_endpoint', '') or ''
).strip()

with ftp_status_lock:
    ftp_status_cache = build_initial_ftp_status_cache(
        app_settings['ftp_targets'])

configure_ping_logger(app_settings['logs_directory'])

class FTPStatusMonitor:
    """Background worker to monitor FTP endpoint availability."""

    def __init__(self):
        self.thread = None
        self.stop_event = threading.Event()
        self.settings = {
            'ftp_targets': sanitize_ftp_targets([]),
            'ftp_ping_interval': DEFAULT_FTP_PING_INTERVAL
        }

    def start(self, settings):
        """Start monitoring with the provided settings."""
        self.stop()

        ftp_targets = sanitize_ftp_targets(settings.get('ftp_targets'))
        ftp_interval = sanitize_ping_interval(
            settings.get('ftp_ping_interval'), DEFAULT_FTP_PING_INTERVAL)

        global ftp_status_cache
        with ftp_status_lock:
            ftp_status_cache = build_initial_ftp_status_cache(ftp_targets)

        self.settings = {
            'ftp_targets': ftp_targets,
            'ftp_ping_interval': ftp_interval
        }

        self.stop_event.clear()
        self.thread = threading.Thread(target=self.run, daemon=True)
        self.thread.start()

    def stop(self):
        """Stop the monitoring thread if it is running."""
        if self.thread and self.thread.is_alive():
            self.stop_event.set()
            self.thread.join(timeout=5)
        self.stop_event.clear()
        self.thread = None

    def run(self):
        """Run the monitoring loop until stopped."""
        while not self.stop_event.is_set():
            try:
                self._poll_once()
            except Exception as exc:  # pragma: no cover - defensive logging
                logger.exception("Unexpected error while polling FTP targets: %s", exc)

            interval = self._get_interval()
            if self.stop_event.wait(interval):
                break

    def _get_interval(self):
        interval = self.settings.get('ftp_ping_interval',
                                      DEFAULT_FTP_PING_INTERVAL)
        try:
            interval = int(interval)
        except (TypeError, ValueError):
            interval = DEFAULT_FTP_PING_INTERVAL

        return max(5, interval)

    def _poll_once(self):
        timestamp = datetime.utcnow().isoformat() + 'Z'
        targets = sanitize_ftp_targets(self.settings.get('ftp_targets'))
        statuses = []

        for index, target in enumerate(targets, start=1):
            host = target['host']
            port = target['port']
            status = 'unconfigured'
            error_message = None

            if host:
                try:
                    with socket.create_connection((host, port), timeout=5):
                        status = 'online'
                except Exception as exc:  # pragma: no cover - network dependent
                    status = 'offline'
                    error_message = str(exc)
                    logger.warning(
                        "FTP status check failed for %s:%s - %s",
                        host, port, exc
                    )

            statuses.append({
                'name': f'FTP Server {index}',
                'host': host,
                'port': port,
                'status': status,
                'error': error_message,
                'last_checked': timestamp
            })

        global ftp_status_cache
        with ftp_status_lock:
            ftp_status_cache = statuses

        self._write_ping_log(statuses, timestamp)
        return statuses

    def _write_ping_log(self, statuses, timestamp):
        """Persist the latest FTP ping results to a rotating log file."""
        if not ping_logger.handlers:
            configure_ping_logger(app_settings.get('logs_directory', 'logs'))
            if not ping_logger.handlers:
                return

        payload = {
            'timestamp': timestamp,
            'results': [
                {
                    'host': status.get('host'),
                    'port': status.get('port'),
                    'status': status.get('status'),
                    'error': status.get('error')
                }
                for status in statuses
            ]
        }

        try:
            ping_logger.info(json.dumps(payload, ensure_ascii=False))
        except Exception as exc:  # pragma: no cover - defensive logging
            logger.exception('Failed to persist ping status log: %s', exc)


    def poll_now(self):
        """Perform a synchronous FTP status check and return the latest result."""
        statuses = self._poll_once()
        return copy.deepcopy(statuses)


class LogParser:
    def __init__(self, logs_dir="logs"):
        self.logs_dir = logs_dir
        self._file_cache = {}
        self._overrides_state = {
            'signature': None,
            'data': {},
            'version': 0
        }
        self._cache_lock = threading.Lock()

    def get_log_files(self):
        """Get all log files sorted by modification time (newest first)"""
        pattern = os.path.join(self.logs_dir, "Transmission.log*")
        log_files = glob.glob(pattern)
        return sorted(log_files, key=os.path.getmtime, reverse=True)

    def _collect_resend_overrides(self, log_files=None):
        """Collect resend overrides from log files with caching."""
        candidate_files = list(log_files or self.get_log_files())
        signature = {}

        for candidate in candidate_files:
            try:
                signature[candidate] = os.path.getmtime(candidate)
            except OSError:
                continue

        with self._cache_lock:
            state = self._overrides_state
            if state['signature'] == signature:
                return dict(state['data']), state['version']

        overrides = {}
        for candidate in candidate_files:
            try:
                with open(candidate, 'r', encoding='utf-8') as override_handle:
                    for line in override_handle:
                        if ('Dashboard-resend-handler' not in line or
                                'resend_result' not in line):
                            continue
                        try:
                            _, json_blob = line.split('resend_result', 1)
                            override_payload = json.loads(json_blob.strip())
                            override_id = (override_payload or {}).get('id_scan')
                            if override_id and override_id not in overrides:
                                overrides[override_id] = override_payload
                        except (ValueError, json.JSONDecodeError):
                            continue
            except IOError:
                continue

        with self._cache_lock:
            previous_version = self._overrides_state['version']
            self._overrides_state = {
                'signature': signature,
                'data': overrides,
                'version': previous_version + 1
            }

        return dict(overrides), previous_version + 1

    def register_resend_override(self, id_scan, *, payload=None, payload_raw=None, log_file=None):
        """Persist resend override details so cached data reflects the latest payload."""
        if not id_scan:
            return

        payload_copy = copy.deepcopy(payload) if isinstance(payload, (dict, list)) else payload
        payload_raw_text = payload_raw if isinstance(payload_raw, str) else None

        with self._cache_lock:
            state = self._overrides_state
            current_override = copy.deepcopy(state['data'].get(id_scan) or {})

            if payload_copy is not None:
                current_override['json_payload'] = payload_copy
            if payload_raw_text is not None:
                current_override['json_payload_raw'] = payload_raw_text
            if log_file:
                current_override['log_file'] = log_file

            state['data'][id_scan] = current_override
            state['version'] += 1

            for cache_entry in self._file_cache.values():
                cache_entry['override_version'] = state['version']
                for item in cache_entry['data']:
                    if str(item.get('id_scan', '')).strip() == str(id_scan):
                        raw = item.setdefault('raw_data', {})
                        if payload_copy is not None:
                            raw['json_payload'] = copy.deepcopy(payload_copy)
                        if payload_raw_text is not None:
                            raw['json_payload_raw'] = payload_raw_text
                        if log_file:
                            item['file_name'] = log_file

    def parse_log_file(self, file_path, global_overrides=None):
        """Parse a single log file and extract JSON data"""
        data = []
        provisional_entries = {}
        completed_task_ids = set()
        known_containers = {}
        known_upload_metadata = {}
        resend_overrides = dict(global_overrides or {})

        container_token_pattern = re.compile(r'^[A-Z0-9\-\+\\/]+$')
        container_separator_pattern = re.compile(r'\s*([+\\/])\s*')
        container_whitespace_pattern = re.compile(r'\s+')

        def normalize_timestamp(value):
            """Return timestamps in the standard '%Y-%m-%d %H:%M:%S' format."""
            if not value:
                return value
            if isinstance(value, datetime):
                return value.strftime('%Y-%m-%d %H:%M:%S')
            if isinstance(value, str):
                text = value.strip()
                if not text:
                    return text
                candidate = text.replace('T', ' ').rstrip('Z').strip()
                for fmt in ('%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d %H:%M:%S'):
                    try:
                        dt_obj = datetime.strptime(candidate, fmt)
                        return dt_obj.strftime('%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        continue
                try:
                    dt_obj = datetime.fromisoformat(candidate)
                    return dt_obj.strftime('%Y-%m-%d %H:%M:%S')
                except ValueError:
                    return candidate
            return value

        def sanitize_container_value(value):
            """Return a normalized container number or empty string if invalid."""
            if value is None:
                return ''
            container_value = str(value).strip()
            if not container_value:
                return ''
            if container_value.lower() == 'failed!':
                return ''
            container_value = container_separator_pattern.sub(r'\1', container_value)
            container_value = container_whitespace_pattern.sub('', container_value)
            container_value = container_value.upper()
            if len(container_value) < 4:
                return ''
            if not any(ch.isdigit() for ch in container_value):
                return ''
            if not any(ch.isalpha() for ch in container_value):
                return ''
            if not container_token_pattern.match(container_value):
                return ''
            return container_value

        def remember_container(entry_id, container_value):
            """Persist sanitized container numbers for reuse across retries."""
            if not entry_id:
                return
            container_value = sanitize_container_value(container_value)
            if not container_value:
                return
            known_containers[entry_id] = container_value

        def sync_entry_container(entry_id, container_value):
            """Update existing entries once a container number is identified."""
            if not entry_id:
                return
            container_value = sanitize_container_value(container_value)
            if not container_value:
                return
            remember_container(entry_id, container_value)
            for existing_entry in reversed(data):
                if existing_entry.get('id_scan') == entry_id:
                    current_container = existing_entry.get('container_no')
                    if (not current_container or
                            str(current_container).strip().lower() == 'failed!'):
                        existing_entry['container_no'] = container_value
                    existing_raw = existing_entry.setdefault('raw_data', {})
                    existing_raw['container_no'] = container_value
                    existing_raw['CONTAINER_NO'] = container_value
                    break

        def remember_upload_metadata(entry_id, metadata):
            """Store supplemental upload metadata for reuse across retries."""
            if not entry_id or metadata is None:
                return
            stored = known_upload_metadata.setdefault(entry_id, {})
            for key, value in metadata.items():
                if value is None:
                    continue
                if isinstance(value, str) and value.strip() in ('', 'N/A'):
                    continue
                if key in ('container_no', 'CONTAINER_NO'):
                    normalized = sanitize_container_value(value)
                    if not normalized:
                        continue
                    stored[key] = normalized
                    continue
                stored[key] = value

        def apply_resend_override(entry_obj):
            """Merge resend override details into the parsed entry."""
            if not entry_obj:
                return
            override = resend_overrides.get(entry_obj.get('id_scan'))
            if not override:
                return

            raw = entry_obj.setdefault('raw_data', {})
            if override.get('json_payload') is not None:
                raw['json_payload'] = copy.deepcopy(override['json_payload'])
            if isinstance(override.get('json_payload_raw'), str) and override['json_payload_raw'].strip():
                raw['json_payload_raw'] = override['json_payload_raw'].strip()
            if override.get('log_file'):
                entry_obj['file_name'] = override['log_file']
            raw['resend_status'] = override.get('status')
            resend_timestamp = normalize_timestamp(override.get('timestamp'))
            raw['resend_timestamp'] = resend_timestamp or override.get('timestamp')
            raw['resend_http_status'] = override.get('http_status')
            response_text = override.get('response_text') or ''
            raw['resend_response_text'] = response_text.replace('\\n', '\n')
            if override.get('target_url'):
                raw.setdefault('post_url', override.get('target_url'))
                raw['resend_target_url'] = override.get('target_url')

            if override.get('status') == 'SUCCESS':
                entry_obj['status'] = 'OK'
                if resend_timestamp:
                    entry_obj['update_time'] = resend_timestamp
                elif override.get('timestamp'):
                    entry_obj['update_time'] = override['timestamp']
                entry_obj['error_description'] = ''
                raw['status'] = 'OK'
            else:
                raw['status'] = entry_obj.get('status')

        def parse_task_time(raw_value):
            """Convert datetime.datetime(...) text to a formatted string."""
            if not raw_value:
                return None
            parts = [part.strip() for part in raw_value.split(',')]
            try:
                if len(parts) < 3:
                    return None
                year = int(parts[0])
                month = int(parts[1])
                day = int(parts[2])
                hour = int(parts[3]) if len(parts) > 3 else 0
                minute = int(parts[4]) if len(parts) > 4 else 0
                second = int(parts[5]) if len(parts) > 5 else 0
                dt_obj = datetime(year, month, day, hour, minute, second)
                return dt_obj.strftime('%Y-%m-%d %H:%M:%S')
            except (ValueError, TypeError, IndexError):
                return None

        def extract_upload_info(line):
            """Extract task details from upload related log lines."""
            timestamp_match = re.search(r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})',
                                        line)
            log_timestamp = timestamp_match.group(1) if timestamp_match else None

            if ('Task.py-send_message_handler' in line and
                    'json_data is' in line):
                decoded_line = line
                try:
                    decoded_line = unescape(line)
                except Exception:
                    pass

                picno_match = re.search(r'<PICNO>([^<]+)</PICNO>', decoded_line, re.IGNORECASE)
                if not picno_match:
                    return None

                scan_time_match = re.search(r'<SCANTIME>([^<]+)</SCANTIME>', decoded_line, re.IGNORECASE)
                container_match = re.search(r'<container_no>([^<]+)</container_no>', decoded_line, re.IGNORECASE)
                checkin_match = re.search(r'<CHECKINTIME>([^<]+)</CHECKINTIME>', decoded_line, re.IGNORECASE)

                info = {
                    'task_no': picno_match.group(1).strip(),
                    'image_path': '',
                    'retry_count': 0,
                    'task_time': scan_time_match.group(1).strip()
                    if scan_time_match else None,
                    'log_timestamp': log_timestamp
                }

                url_match = re.search(r'url is\s*([^,]+)', decoded_line, re.IGNORECASE)
                if url_match:
                    info['post_url'] = url_match.group(1).strip()

                json_match = re.search(r'json_data is (.*)$', decoded_line, re.IGNORECASE)
                if json_match:
                    json_raw = json_match.group(1).strip()
                    json_raw = normalize_container_fields_in_text(json_raw, escape_for_literal=True)
                    info['json_payload_raw'] = json_raw
                    try:
                        parsed_payload = ast.literal_eval(json_raw)
                        normalize_containers_in_payload(parsed_payload)
                        info['json_payload'] = parsed_payload
                    except (ValueError, SyntaxError):
                        pass

                if container_match:
                    container_value = sanitize_container_value(container_match.group(1))
                    if container_value:
                        info['container_no'] = container_value
                        remember_container(info['task_no'], container_value)

                if checkin_match:
                    update_time_value = checkin_match.group(1).strip()
                    if update_time_value:
                        info['update_time'] = update_time_value

                scanimg_count = len(re.findall(r'<SCANIMG', decoded_line, re.IGNORECASE))
                img_tag_count = len(re.findall(r'<img>', decoded_line, re.IGNORECASE))
                image_count = max(scanimg_count, img_tag_count)
                if image_count:
                    info['image_count'] = image_count

                scan_start_match = (
                    re.search(r'<Time_ScanStart>([^<]+)</Time_ScanStart>', decoded_line, re.IGNORECASE) or
                    re.search(r'<Time_Scan_Start>([^<]+)</Time_Scan_Start>', decoded_line, re.IGNORECASE)
                )
                scan_stop_match = (
                    re.search(r'<Time_Scan_Stop>([^<]+)</Time_Scan_Stop>', decoded_line, re.IGNORECASE) or
                    re.search(r'<Time_ScanStop>([^<]+)</Time_ScanStop>', decoded_line, re.IGNORECASE)
                )
                scan_duration_value = None
                if scan_start_match and scan_stop_match:
                    try:
                        start_val = int(scan_start_match.group(1).strip())
                        stop_val = int(scan_stop_match.group(1).strip())
                        if stop_val >= start_val:
                            seconds = stop_val - start_val
                            scan_duration_value = f"{seconds} detik"
                    except ValueError:
                        pass

                if scan_duration_value:
                    info['scan_duration'] = scan_duration_value
                    info['overall_time'] = scan_duration_value

                scan_time_value = info.get('task_time')
                update_time_candidate = info.get('update_time') or log_timestamp
                if scan_time_value and update_time_candidate:
                    try:
                        time_diff_value = self.calculate_time_difference({
                            'SCANTIME': scan_time_value,
                            'UPDATE_TIME': update_time_candidate
                        })
                        if time_diff_value and time_diff_value != 'N/A':
                            info['time_difference'] = time_diff_value
                    except Exception:
                        pass

                metadata_snapshot = {
                    'task_time': info.get('task_time'),
                    'update_time': info.get('update_time'),
                    'scan_duration': info.get('scan_duration'),
                    'overall_time': info.get('overall_time'),
                    'time_difference': info.get('time_difference'),
                    'image_count': info.get('image_count'),
                    'container_no': info.get('container_no'),
                    'log_timestamp': log_timestamp
                }
                remember_upload_metadata(info['task_no'], metadata_snapshot)

                return info

            if ('Task.py-build_upload_data' not in line and
                    'XmlParse.py-parse_xml' not in line):
                return None

            task_no_match = (re.search(r"'task_no':\s*'([^']+)'", line) or
                             re.search(r"'pic_no':\s*'([^']+)'", line))
            image_path_match = (re.search(r"'image_path':\s*'([^']*)'", line) or
                                re.search(r"'img_dir_path':\s*'([^']*)'", line))
            retry_match = re.search(r"'retry_(?:count|time)':\s*(\d+)", line)
            task_time_match = re.search(
                r"'task_time':\s*datetime\.datetime\(([^)]+)\)", line)

            task_time = parse_task_time(task_time_match.group(1)
                                        if task_time_match else None)

            if not task_no_match:
                return None

            return {
                'task_no': task_no_match.group(1),
                'image_path': image_path_match.group(1)
                if image_path_match else '',
                'retry_count': int(retry_match.group(1))
                if retry_match else 0,
                'task_time': task_time,
                'log_timestamp': log_timestamp
            }

        def update_provisional_entry(info):
            """Create or update provisional entries for upload events."""
            task_no = info.get('task_no')
            if not task_no or task_no in completed_task_ids:
                return

            log_timestamp = info.get('log_timestamp')
            scan_time = info.get('task_time') or log_timestamp or 'N/A'
            image_path = info.get('image_path', '')
            retry_count = info.get('retry_count', 0)
            container_no = sanitize_container_value(info.get('container_no'))
            scan_duration = info.get('scan_duration')
            overall_time = info.get('overall_time') or scan_duration
            time_difference = info.get('time_difference')
            update_time_value = info.get('update_time') or log_timestamp
            image_count_value = info.get('image_count')
            if isinstance(image_count_value, str):
                stripped_count = image_count_value.strip()
                if stripped_count.isdigit():
                    image_count_value = int(stripped_count)
            elif image_count_value is not None:
                try:
                    image_count_value = int(image_count_value)
                except (TypeError, ValueError):
                    pass

            effective_scan_duration = scan_duration if scan_duration else 'N/A'
            effective_overall_time = overall_time or (scan_duration if scan_duration else None)
            if not effective_overall_time:
                effective_overall_time = 'N/A'
            effective_time_difference = time_difference if time_difference else 'N/A'
            if ((not effective_time_difference) or effective_time_difference == 'N/A') and scan_time and update_time_value:
                try:
                    computed_diff = self.calculate_time_difference({
                        'SCANTIME': scan_time,
                        'UPDATE_TIME': update_time_value
                    })
                    if computed_diff and computed_diff != 'N/A':
                        effective_time_difference = computed_diff
                except Exception:
                    pass

            entry = provisional_entries.get(task_no)
            if entry:
                raw_data = entry.setdefault('raw_data', {})
                raw_data['task_no'] = task_no

                post_url_value = info.get('post_url')
                if post_url_value:
                    raw_data['post_url'] = post_url_value

                if scan_time and scan_time != 'N/A':
                    entry['scan_time'] = scan_time
                    raw_data['task_time'] = scan_time

                if update_time_value:
                    entry['update_time'] = update_time_value
                    raw_data['update_time'] = update_time_value
                elif 'update_time' not in raw_data and log_timestamp:
                    raw_data['update_time'] = log_timestamp

                if log_timestamp:
                    entry['log_timestamp'] = log_timestamp
                    raw_data['log_timestamp'] = log_timestamp

                if container_no:
                    entry['container_no'] = container_no
                    raw_data['container_no'] = container_no
                    raw_data['CONTAINER_NO'] = container_no
                    remember_container(task_no, container_no)
                    sync_entry_container(task_no, container_no)

                if scan_duration and scan_duration != 'N/A':
                    entry['scan_duration'] = scan_duration
                    entry['overall_time'] = overall_time or scan_duration
                    raw_data['scan_duration'] = scan_duration
                    if overall_time:
                        raw_data['overall_time'] = overall_time
                elif overall_time and overall_time != 'N/A':
                    entry['overall_time'] = overall_time
                    raw_data['overall_time'] = overall_time

                if effective_time_difference and effective_time_difference != 'N/A':
                    entry['time_difference'] = effective_time_difference
                    raw_data['time_difference'] = effective_time_difference
                elif time_difference and time_difference != 'N/A':
                    entry['time_difference'] = time_difference
                    raw_data['time_difference'] = time_difference

                if image_count_value is not None:
                    try:
                        count_int = int(image_count_value)
                    except (TypeError, ValueError):
                        count_int = image_count_value
                    entry['image_count'] = count_int
                    raw_data['image_count'] = count_int

                if image_path:
                    raw_data['image_path'] = image_path

                if retry_count or 'retry_count' not in raw_data:
                    raw_data['retry_count'] = retry_count

                if 'json_payload' in info and info['json_payload'] is not None:
                    raw_data['json_payload'] = info['json_payload']

                if 'json_payload_raw' in info and info['json_payload_raw']:
                    raw_data['json_payload_raw'] = info['json_payload_raw']

                remember_upload_metadata(task_no, {
                    'task_time': scan_time if scan_time != 'N/A' else None,
                    'update_time': update_time_value,
                    'scan_duration': effective_scan_duration if effective_scan_duration != 'N/A' else None,
                    'overall_time': effective_overall_time if effective_overall_time != 'N/A' else None,
                    'time_difference': entry.get('time_difference'),
                    'image_count': entry.get('image_count'),
                    'container_no': container_no or None,
                    'log_timestamp': log_timestamp,
                    'post_url': post_url_value,
                    'json_payload': info.get('json_payload') if info.get('json_payload') is not None else None,
                    'json_payload_raw': info.get('json_payload_raw')
                })

                return

            raw_data = {
                'task_no': task_no,
                'image_path': image_path,
                'retry_count': retry_count,
                'task_time': scan_time,
                'source': 'upload'
            }
            if update_time_value:
                raw_data['update_time'] = update_time_value
            if image_count_value is not None:
                raw_data['image_count'] = image_count_value
            if log_timestamp:
                raw_data['log_timestamp'] = log_timestamp
            if container_no:
                raw_data['container_no'] = container_no
                raw_data['CONTAINER_NO'] = container_no
                remember_container(task_no, container_no)
            if scan_duration and scan_duration != 'N/A':
                raw_data['scan_duration'] = scan_duration
            if overall_time and overall_time != 'N/A':
                raw_data['overall_time'] = overall_time
            if effective_time_difference and effective_time_difference != 'N/A':
                raw_data['time_difference'] = effective_time_difference
            elif time_difference and time_difference != 'N/A':
                raw_data['time_difference'] = time_difference

            post_url_value = info.get('post_url')
            if post_url_value:
                raw_data['post_url'] = post_url_value

            if 'json_payload' in info and info['json_payload'] is not None:
                raw_data['json_payload'] = info['json_payload']

            if 'json_payload_raw' in info:
                raw_data['json_payload_raw'] = info['json_payload_raw']

            provisional_entries[task_no] = {
                'id_scan': task_no,
                'container_no': container_no or '',
                'scan_time': scan_time,
                'scan_duration': effective_scan_duration,
                'overall_time': effective_overall_time,
                'update_time': update_time_value or (log_timestamp or 'N/A'),
                'time_difference': effective_time_difference,
                'image_count': image_count_value if image_count_value is not None else 0,
                'status': 'NOK',
                'log_timestamp': log_timestamp,
                'file_name': os.path.basename(file_path),
                'raw_data': raw_data
            }

            remember_upload_metadata(task_no, {
                'task_time': scan_time if scan_time != 'N/A' else None,
                'update_time': update_time_value or log_timestamp,
                'scan_duration': effective_scan_duration if effective_scan_duration != 'N/A' else None,
                'overall_time': effective_overall_time if effective_overall_time != 'N/A' else None,
                'time_difference': effective_time_difference if effective_time_difference != 'N/A' else None,
                'image_count': image_count_value if image_count_value is not None else 0,
                'container_no': container_no or None,
                'log_timestamp': log_timestamp,
                'post_url': post_url_value,
                'json_payload': info.get('json_payload') if info.get('json_payload') is not None else None,
                'json_payload_raw': info.get('json_payload_raw')
            })

            sync_entry_container(task_no, container_no)

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                for line in f:
                    if ('Dashboard-resend-handler' in line and
                            'resend_result' in line):
                        try:
                            _, json_blob = line.split('resend_result', 1)
                            override_payload = json.loads(json_blob.strip())
                            override_id = (override_payload or {}).get('id_scan')
                            if override_id:
                                resend_overrides[override_id] = override_payload
                        except (ValueError, json.JSONDecodeError):
                            pass
                        continue

                    upload_info = extract_upload_info(line)
                    if upload_info:
                        update_provisional_entry(upload_info)
                        continue

                    # Look for lines with response text (both success and failure)
                    if 'response text:' in line and 'center response:' in line:
                        try:
                            # Extract the ID from center response
                            id_match = re.search(r'center response:([^,]+)', line)
                            response_id = id_match.group(1) if id_match else ''
                            
                            # Extract JSON from the line
                            json_start = line.find('response text: ') + len('response text: ')
                            json_str = line[json_start:].strip()
                            
                            # Parse the JSON
                            response_data = json.loads(json_str)
                            
                            # Extract timestamp from log line
                            pattern = r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})'
                            timestamp_match = re.search(pattern, line)
                            log_timestamp = (timestamp_match.group(1) 
                                           if timestamp_match else None)
                            
                            # Handle successful responses (resultCode: true)
                            if (response_data.get('resultCode') and 
                                    response_data.get('resultData') and
                                    response_data.get('resultData') != '-'):
                                result_data = response_data['resultData']

                                container_no = sanitize_container_value(result_data.get('CONTAINER_NO', ''))
                                provisional_entry = None
                                result_picno = result_data.get('PICNO')
                                if result_picno:
                                    provisional_entry = provisional_entries.get(result_picno)
                                upload_meta = known_upload_metadata.get(result_picno, {}) if result_picno else {}
                                if provisional_entry:
                                    provisional_raw = provisional_entry.get('raw_data', {})
                                    for key, value in provisional_raw.items():
                                        result_data.setdefault(key, value)
                                for key, value in upload_meta.items():
                                    result_data.setdefault(key, value)
                                xml_container = None
                                if provisional_entry:
                                    xml_container = (provisional_entry.get('container_no') or
                                                     provisional_entry.get('raw_data', {}).get('container_no') or
                                                     provisional_entry.get('raw_data', {}).get('CONTAINER_NO'))
                                clean_container = container_no
                                if not clean_container and xml_container:
                                    clean_container = sanitize_container_value(xml_container)
                                if not clean_container and result_picno:
                                    known_value = known_containers.get(result_picno)
                                    if known_value:
                                        clean_container = known_value
                                if not clean_container and upload_meta.get('container_no'):
                                    clean_container = sanitize_container_value(upload_meta['container_no'])
                                container_no = clean_container
                                if container_no:
                                    result_data['CONTAINER_NO'] = container_no
                                    result_data['container_no'] = container_no
                                    remember_container(result_picno, container_no)
                                sync_entry_container(result_picno, container_no)

                                
                                # Calculate scan duration
                                scan_duration = self.calculate_scan_duration(
                                    result_data)
                                if (not scan_duration or scan_duration == 'N/A') and upload_meta.get('scan_duration'):
                                    scan_duration = upload_meta['scan_duration']

                                # Calculate time difference
                                time_diff = self.calculate_time_difference(
                                    result_data)
                                if (not time_diff or time_diff == 'N/A') and upload_meta.get('time_difference'):
                                    time_diff = upload_meta['time_difference']

                                # Count images
                                image_count = self.count_images(result_data)
                                if (not image_count or image_count == 0) and upload_meta.get('image_count') is not None:
                                    try:
                                        image_count = int(upload_meta['image_count'])
                                    except (TypeError, ValueError):
                                        image_count = upload_meta['image_count']
                                
                                # Determine status
                                status = ("OK" if result_data.get('RESPON_TPS_API') 
                                         == 'OK' else "NOK")
                                
                                if not result_data.get('UPDATE_TIME') and (upload_meta.get('update_time') or upload_meta.get('log_timestamp') or log_timestamp):
                                    result_data['UPDATE_TIME'] = upload_meta.get('update_time') or upload_meta.get('log_timestamp') or log_timestamp
                                effective_scan_time = result_data.get('SCANTIME') or upload_meta.get('task_time') or log_timestamp or 'N/A'
                                effective_update_time = result_data.get('UPDATE_TIME') or upload_meta.get('update_time') or upload_meta.get('log_timestamp') or log_timestamp or 'N/A'
                                overall_time_value = scan_duration if scan_duration and scan_duration != 'N/A' else upload_meta.get('overall_time', scan_duration)
                                if image_count is None:
                                    image_count = upload_meta.get('image_count') if upload_meta.get('image_count') is not None else 0
                                result_data.setdefault('image_count', image_count)
                                if effective_update_time and effective_update_time != 'N/A':
                                    result_data['UPDATE_TIME'] = effective_update_time
                                effective_container = container_no or known_containers.get(result_data.get('PICNO', '')) or ''
                                entry = {
                                    'id_scan': result_data.get('PICNO', ''),
                                    'container_no': effective_container,
                                    'scan_time': effective_scan_time,
                                    'scan_duration': scan_duration,
                                    'overall_time': overall_time_value,
                                    'update_time': effective_update_time,
                                    'time_difference': time_diff,
                                    'image_count': image_count,
                                    'status': status,
                                    'log_timestamp': log_timestamp,
                                    'file_name': os.path.basename(file_path),
                                    'raw_data': result_data
                                }
                                data.append(entry)
                                if entry['id_scan']:
                                    completed_task_ids.add(entry['id_scan'])
                                    provisional_entries.pop(entry['id_scan'], None)
                                elif response_id:
                                    completed_task_ids.add(response_id)
                                    provisional_entries.pop(response_id, None)

                            # Handle failed responses (resultCode: false)
                            elif (response_data.get('resultCode') is False and
                                  response_data.get('resultData') == '-'):

                                # Extract container number from response description if available
                                container_no = ""
                                desc = response_data.get('resultDesc', '')
                                if 'Container' in desc:
                                    # Try to extract container number from description
                                    container_match = re.search(
                                        r'Container[^:]*:?\s*([A-Z0-9]+)', desc)
                                    if container_match:
                                        container_no = sanitize_container_value(container_match.group(1))

                                provisional_entry = provisional_entries.get(response_id) if response_id else None
                                upload_meta = known_upload_metadata.get(response_id, {}) if response_id else {}
                                xml_container = None
                                if provisional_entry:
                                    raw_provisional = provisional_entry.get('raw_data', {})
                                    xml_container = (provisional_entry.get('container_no') or
                                                     raw_provisional.get('container_no') or
                                                     raw_provisional.get('CONTAINER_NO'))
                                if xml_container:
                                    container_no = sanitize_container_value(xml_container) or container_no
                                if not container_no and upload_meta.get('container_no'):
                                    container_no = sanitize_container_value(upload_meta['container_no'])

                                if not container_no and response_id:
                                    known_value = known_containers.get(response_id)
                                    if known_value:
                                        container_no = known_value
                                if container_no:
                                    remember_container(response_id, container_no)
                                sync_entry_container(response_id, container_no)

                                def resolve_field(field_name, fallback='N/A'):
                                    if not provisional_entry:
                                        return fallback
                                    value = provisional_entry.get(field_name)
                                    if not value or value == 'N/A':
                                        value = provisional_entry.get('raw_data', {}).get(field_name)
                                    if not value or value == 'N/A':
                                        return fallback
                                    return value

                                scan_duration_value = resolve_field('scan_duration')
                                overall_time_value = resolve_field('overall_time')
                                if overall_time_value == 'N/A' and scan_duration_value != 'N/A':
                                    overall_time_value = scan_duration_value
                                if (not scan_duration_value or scan_duration_value == 'N/A') and upload_meta.get('scan_duration'):
                                    scan_duration_value = upload_meta['scan_duration']
                                if (not overall_time_value or overall_time_value == 'N/A') and upload_meta.get('overall_time'):
                                    overall_time_value = upload_meta['overall_time']
                                time_difference_value = resolve_field('time_difference')
                                if (not time_difference_value or time_difference_value == 'N/A') and upload_meta.get('time_difference'):
                                    time_difference_value = upload_meta['time_difference']

                                scan_time_value = log_timestamp or 'N/A'
                                if (provisional_entry and provisional_entry.get('scan_time') and
                                        provisional_entry['scan_time'] != 'N/A'):
                                    scan_time_value = provisional_entry['scan_time']
                                elif provisional_entry:
                                    raw_task_time = provisional_entry.get('raw_data', {}).get('task_time')
                                    if raw_task_time:
                                        scan_time_value = raw_task_time
                                if (not scan_time_value or scan_time_value == 'N/A') and upload_meta.get('task_time'):
                                    scan_time_value = upload_meta['task_time']

                                entry_raw_data = dict(response_data)
                                if provisional_entry:
                                    provisional_raw = provisional_entry.get('raw_data', {})
                                    for key, value in provisional_raw.items():
                                        entry_raw_data.setdefault(key, value)
                                for key, value in upload_meta.items():
                                    entry_raw_data.setdefault(key, value)
                                if container_no and container_no.lower() != 'failed!':
                                    entry_raw_data['container_no'] = container_no
                                    entry_raw_data['CONTAINER_NO'] = container_no
                                elif response_id:
                                    known_value = known_containers.get(response_id)
                                    if known_value:
                                        entry_raw_data['container_no'] = known_value
                                        entry_raw_data['CONTAINER_NO'] = known_value
                                if scan_duration_value != 'N/A':
                                    entry_raw_data['scan_duration'] = scan_duration_value
                                if overall_time_value != 'N/A':
                                    entry_raw_data['overall_time'] = overall_time_value
                                if time_difference_value != 'N/A':
                                    entry_raw_data['time_difference'] = time_difference_value
                                effective_update_time = upload_meta.get('update_time') or upload_meta.get('log_timestamp') or log_timestamp or 'N/A'
                                if effective_update_time and effective_update_time != 'N/A':
                                    entry_raw_data.setdefault('update_time', effective_update_time)

                                effective_container = sanitize_container_value(container_no) or known_containers.get(response_id, '')
                                image_count_effective = entry_raw_data.get('image_count')
                                if image_count_effective is None and provisional_entry:
                                    image_count_effective = (provisional_entry.get('image_count') or
                                                             provisional_entry.get('raw_data', {}).get('image_count'))
                                if image_count_effective is None:
                                    image_count_effective = 0
                                try:
                                    image_count_effective = int(image_count_effective)
                                except (TypeError, ValueError):
                                    pass

                                entry = {
                                    'id_scan': response_id,
                                    'container_no': effective_container,
                                    'scan_time': scan_time_value,
                                    'scan_duration': scan_duration_value,
                                    'overall_time': overall_time_value,
                                    'update_time': effective_update_time,
                                    'time_difference': time_difference_value,
                                    'image_count': image_count_effective,
                                    'status': 'NOK',
                                    'log_timestamp': log_timestamp,
                                    'file_name': os.path.basename(file_path),
                                    'raw_data': entry_raw_data,
                                    'error_description': desc
                                }
                                data.append(entry)
                                if response_id:
                                    completed_task_ids.add(response_id)
                                    provisional_entries.pop(response_id, None)

                        except (json.JSONDecodeError, KeyError):
                            continue
        except IOError as e:
            print(f"Error reading file {file_path}: {e}")

        for entry in data:
            apply_resend_override(entry)

        for provisional in provisional_entries.values():
            apply_resend_override(provisional)

        data.extend(provisional_entries.values())

        return data
    
    def calculate_scan_duration(self, data):
        """Calculate scan duration from TIME_SCANSTART and TIME_SCAN_STOP"""
        try:
            start_time = data.get('TIME_SCANSTART', 0)
            stop_time = data.get('TIME_SCAN_STOP', 0)
            if start_time and stop_time:
                duration = int(stop_time) - int(start_time)
                return f"{duration} detik"
        except (ValueError, TypeError):
            pass
        return "N/A"
    
    def calculate_time_difference(self, data):
        """Calculate time difference between UPDATE_TIME and SCANTIME"""
        try:
            scan_time = data.get('SCANTIME', '')
            update_time = data.get('UPDATE_TIME', '')
            
            if scan_time and update_time:
                scan_dt = datetime.strptime(scan_time, '%Y-%m-%d %H:%M:%S')
                update_dt = datetime.strptime(update_time, '%Y-%m-%d %H:%M:%S')
                diff = update_dt - scan_dt
                
                if diff.days < 0:
                    hours = diff.seconds // 3600
                    minutes = (diff.seconds % 3600) // 60
                    seconds = diff.seconds % 60
                    return (f"-{abs(diff.days)} day, "
                            f"{hours:02d}:{minutes:02d}:{seconds:02d}")
                else:
                    hours = diff.seconds // 3600
                    minutes = (diff.seconds % 3600) // 60
                    seconds = diff.seconds % 60
                    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        except (ValueError, TypeError):
            pass
        return "N/A"
    
    def count_images(self, data):
        """Count the number of images"""
        count = 0
        for i in range(1, 8):  # IMAGE1_PATH to IMAGE7_PATH
            if data.get(f'IMAGE{i}_PATH'):
                count += 1
        return count
    
    def get_all_data(self, status_filter=None, search_term=None,
                     log_file=None):
        """Get all data from all log files with optional filtering"""
        all_data = []

        all_log_files = self.get_log_files()
        global_resend_overrides, overrides_version = (
            self._collect_resend_overrides(all_log_files))

        # Remove cache entries for files that no longer exist
        existing_paths = set(all_log_files)
        with self._cache_lock:
            cached_paths = list(self._file_cache.keys())
            for cached_path in cached_paths:
                if cached_path not in existing_paths:
                    self._file_cache.pop(cached_path, None)

        # Filter by specific log file if specified
        if log_file:
            log_files = [
                f for f in all_log_files
                if os.path.basename(f) == log_file
            ]
        else:
            log_files = all_log_files

        for file_path in log_files:
            try:
                file_mtime = os.path.getmtime(file_path)
            except OSError:
                continue

            with self._cache_lock:
                cache_entry = self._file_cache.get(file_path)

            if (cache_entry and cache_entry['mtime'] == file_mtime and
                    cache_entry['override_version'] == overrides_version):
                file_data = cache_entry['data']
            else:
                file_data = self.parse_log_file(file_path, global_resend_overrides)
                with self._cache_lock:
                    self._file_cache[file_path] = {
                        'mtime': file_mtime,
                        'override_version': overrides_version,
                        'data': file_data
                    }

            all_data.extend(file_data)
        
        # Remove duplicates based on ID scan (keep the latest one)
        seen_ids = set()
        unique_data = []
        
        # Sort by scan time first to ensure we keep the latest entry
        all_data.sort(key=lambda x: x['scan_time'], reverse=True)
        
        for entry in all_data:
            if entry['id_scan'] and entry['id_scan'] not in seen_ids:
                seen_ids.add(entry['id_scan'])
                unique_data.append(entry)
        
        # Apply filters after deduplication
        if status_filter:
            unique_data = [entry for entry in unique_data 
                          if entry['status'] == status_filter]
        
        if search_term:
            search_term = search_term.lower()
            unique_data = [entry for entry in unique_data if 
                          search_term in entry['id_scan'].lower() or 
                          search_term in entry['container_no'].lower()]
        
        return unique_data

    def find_json_payload(self, task_no, log_file=None):
        """Locate the original JSON payload for a given task by scanning the logs."""
        if not task_no:
            return None

        log_files = self.get_log_files()
        if log_file:
            log_files = [
                f for f in log_files
                if os.path.basename(f) == os.path.basename(log_file)
            ]

        for file_path in log_files:
            try:
                with open(file_path, 'r', encoding='utf-8') as handle:
                    for line in handle:
                        if ('Task.py-send_message_handler' not in line or
                                'json_data is' not in line or
                                task_no not in line):
                            continue

                        decoded_line = unescape(line)
                        url_match = re.search(r'url is\s*([^,]+)', decoded_line, re.IGNORECASE)
                        json_match = re.search(r'json_data is (.*)$', decoded_line, re.IGNORECASE)
                        if not json_match:
                            continue

                        raw_payload = json_match.group(1).strip()
                        raw_payload = normalize_container_fields_in_text(raw_payload, escape_for_literal=True)
                        parsed_payload = None
                        try:
                            parsed_payload = ast.literal_eval(raw_payload)
                            normalize_containers_in_payload(parsed_payload)
                        except (ValueError, SyntaxError):
                            try:
                                parsed_payload = json.loads(raw_payload)
                                normalize_containers_in_payload(parsed_payload)
                            except json.JSONDecodeError:
                                parsed_payload = None

                        return {
                            'post_url': url_match.group(1).strip() if url_match else None,
                            'payload': parsed_payload,
                            'payload_raw': raw_payload
                        }
            except IOError:
                continue

        return None

# Initialize log parser with settings
log_parser = LogParser(app_settings['logs_directory'])

# Start FTP monitoring thread
ftp_monitor = FTPStatusMonitor()
ftp_monitor.start(app_settings)


@app.route('/')
def dashboard():
    """Main dashboard page"""
    return render_template('dashboard.html')


@app.route('/api/data')
def get_data():
    """API endpoint to get log data"""
    status_filter = request.args.get('status')
    search_term = request.args.get('search')
    log_file = request.args.get('log_file')
    
    data = log_parser.get_all_data(status_filter, search_term, log_file)
    
    return jsonify({
        'data': data,
        'total': len(data)
    })


@app.route('/api/resend', methods=['POST'])
def resend_payload():
    """API endpoint to resend payload data for a specific scan."""
    request_payload = request.get_json(silent=True) or {}
    id_scan = str(request_payload.get('id_scan') or '').strip()
    log_file = request_payload.get('log_file')
    payload_override = request_payload.get('payload_override')
    payload_override_raw = request_payload.get('payload_override_raw')

    if not id_scan:
        return jsonify({'error': 'id_scan is required'}), 400

    if log_file:
        log_file = str(log_file).strip()
        if not log_file:
            log_file = None

    server_value = app_settings.get('resend_server', '')
    endpoint_value = app_settings.get('resend_endpoint', '')

    try:
        target_url = build_resend_url(server_value, endpoint_value)
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400

    try:
        entries = (log_parser.get_all_data(log_file=log_file)
                   if log_file else log_parser.get_all_data())
    except Exception as exc:  # pragma: no cover - defensive logging
        logger.exception("Failed to load log data for resend request: %s", exc)
        return jsonify({'error': 'Failed to load log data'}), 500

    entry = next(
        (item for item in entries
         if str(item.get('id_scan', '')).strip() == id_scan),
        None
    )

    if not entry:
        return jsonify({'error': f'ID scan {id_scan} was not found'}), 404

    def log_resend_outcome(entry_data, log_file_hint, status_value, response_obj=None, response_text_value='', target_url_value=None):
        """Append resend outcome details (success or failure) to the transmission log."""
        try:
            logs_dir = app_settings.get('logs_directory', 'logs')
        except Exception:
            logs_dir = 'logs'

        log_filename = 'Transmission.log'
        log_path = os.path.join(logs_dir, log_filename)

        timestamp = datetime.utcnow()
        line_timestamp = timestamp.strftime('%Y-%m-%d %H:%M:%S,%f')[:-3]

        payload_timestamp = timestamp.strftime('%Y-%m-%d %H:%M:%S')

        payload = {
            'id_scan': entry_data.get('id_scan'),
            'status': status_value,
            'http_status': getattr(response_obj, 'status_code', None),
            'target_url': target_url_value,
            'response_text': response_text_value,
            'log_file': log_filename,
            'timestamp': payload_timestamp
        }

        try:
            os.makedirs(os.path.dirname(log_path), exist_ok=True)
        except OSError:
            pass

        try:
            with open(log_path, 'a', encoding='utf-8') as handle:
                handle.write(
                    f"{line_timestamp} INFO [Dashboard-resend-handler] resend_result "
                    f"{json.dumps(payload, ensure_ascii=False)}\n"
                )
        except OSError:
            logger.exception("Failed to append resend outcome to %s", log_path)

    raw_data = entry.get('raw_data') or {}
    json_payload = raw_data.get('json_payload')
    payload_raw = raw_data.get('json_payload_raw')
    post_url = raw_data.get('post_url')

    def coerce_payload(value):
        if isinstance(value, (dict, list)):
            return value
        if isinstance(value, str):
            try:
                return json.loads(value)
            except json.JSONDecodeError:
                try:
                    return ast.literal_eval(value)
                except (ValueError, SyntaxError):
                    return None
        return None

    override_applied = False

    if isinstance(payload_override, str) and payload_override.strip():
        payload_override = coerce_payload(payload_override)

    if isinstance(payload_override, (dict, list)):
        json_payload = payload_override
        try:
            payload_raw = json.dumps(payload_override, ensure_ascii=False)
        except (TypeError, ValueError):
            payload_raw = None
        override_applied = True
    elif isinstance(payload_override_raw, str) and payload_override_raw.strip():
        payload_raw = payload_override_raw.strip()
        candidate_override = coerce_payload(payload_raw)
        if isinstance(candidate_override, (dict, list)):
            json_payload = candidate_override
        override_applied = True

    if override_applied:
        if isinstance(json_payload, (dict, list)):
            raw_data['json_payload'] = json_payload
        if isinstance(payload_raw, str):
            raw_data['json_payload_raw'] = payload_raw

    if isinstance(json_payload, str):
        json_payload = coerce_payload(json_payload)

    if not isinstance(json_payload, (dict, list)):
        json_payload = coerce_payload(payload_raw)

    fallback_payload = None
    if not isinstance(json_payload, (dict, list)) or not post_url:
        log_file_hint = log_file or entry.get('file_name')
        fallback_payload = log_parser.find_json_payload(id_scan, log_file_hint)
        if fallback_payload:
            if not isinstance(json_payload, (dict, list)):
                json_payload = fallback_payload.get('payload')
            if not payload_raw:
                payload_raw = fallback_payload.get('payload_raw')
            if not post_url:
                post_url = fallback_payload.get('post_url')
            if payload_raw:
                raw_data.setdefault('json_payload_raw', payload_raw)
            if post_url:
                raw_data['post_url'] = post_url

    if not isinstance(json_payload, (dict, list)):
        json_payload = coerce_payload(payload_raw)

    if isinstance(json_payload, (dict, list)):
        normalize_containers_in_payload(json_payload)
        raw_data['json_payload'] = json_payload
        if isinstance(payload_raw, str) and payload_raw:
            payload_raw = normalize_container_fields_in_text(payload_raw, escape_for_literal=True)
            raw_data['json_payload_raw'] = payload_raw
    elif payload_raw:
        if isinstance(payload_raw, (bytes, bytearray)):
            try:
                payload_raw = payload_raw.decode('utf-8')
            except UnicodeDecodeError:
                payload_raw = payload_raw.decode('latin-1', errors='ignore')
        payload_raw = payload_raw.strip() if isinstance(payload_raw, str) else payload_raw
        if isinstance(payload_raw, str) and payload_raw:
            payload_raw = normalize_container_fields_in_text(payload_raw, escape_for_literal=True)
            raw_data['json_payload_raw'] = payload_raw
        else:
            payload_raw = None

    if not isinstance(json_payload, (dict, list)) and not isinstance(payload_raw, str):
        return jsonify({
            'error': 'No resend payload is available for this entry'
        }), 400

    if fallback_payload and fallback_payload.get('post_url'):
        target_url = fallback_payload['post_url']

    if not post_url:
        post_url = None

    try:
        request_kwargs = {
            'timeout': DEFAULT_RESEND_TIMEOUT
        }
        if isinstance(json_payload, (dict, list)):
            request_kwargs['json'] = json_payload
        else:
            request_kwargs['data'] = payload_raw
            request_kwargs['headers'] = {
                'Content-Type': 'application/json'
            }

        response = requests.post(
            target_url,
            **request_kwargs
        )
    except requests.RequestException as exc:
        logger.exception("Failed to resend data for %s: %s", id_scan, exc)
        log_resend_outcome(
            entry,
            log_file,
            'FAILED',
            response_obj=None,
            response_text_value=str(exc),
            target_url_value=target_url
        )
        return jsonify({'error': f'Failed to send data: {exc}'}), 500

    full_response_text = response.text or ''
    response_text_preview = full_response_text
    if len(response_text_preview) > MAX_REMOTE_RESPONSE_PREVIEW:
        response_text_preview = response_text_preview[:MAX_REMOTE_RESPONSE_PREVIEW] + '...'

    resend_success = interpret_resend_response_success(response, full_response_text)
    outcome_status = 'SUCCESS' if resend_success else 'FAILED'
    log_resend_outcome(
        entry,
        log_file,
        outcome_status,
        response_obj=response,
        response_text_value=response_text_preview.replace('\n', '\\n'),
        target_url_value=target_url
    )

    if override_applied and resend_success:
        log_parser.register_resend_override(
            id_scan,
            payload=json_payload if isinstance(json_payload, (dict, list)) else None,
            payload_raw=payload_raw if isinstance(payload_raw, str) else None,
            log_file=log_file or entry.get('file_name')
        )

    return jsonify({
        'success': resend_success,
        'status_code': response.status_code,
        'response_text': response_text_preview,
        'target_url': target_url
    })


@app.route('/api/log-files')
def get_log_files():
    """API endpoint to get available log files"""
    log_files = log_parser.get_log_files()
    file_list = [os.path.basename(f) for f in log_files]
    return jsonify(file_list)


@app.route('/api/stats')
def get_stats():
    """API endpoint to get statistics"""
    all_data = log_parser.get_all_data()
    
    total_scans = len(all_data)
    ok_scans = len([d for d in all_data if d['status'] == 'OK'])
    nok_scans = len([d for d in all_data if d['status'] == 'NOK'])
    
    # Get recent scans (last 24 hours)
    recent_scans = []
    now = datetime.now()
    for entry in all_data:
        try:
            scan_time = datetime.strptime(entry['scan_time'], 
                                        '%Y-%m-%d %H:%M:%S')
            if (now - scan_time).total_seconds() < 86400:  # 24 hours
                recent_scans.append(entry)
        except (ValueError, TypeError):
            continue
    
    success_rate = (round((ok_scans / total_scans * 100), 2) 
                   if total_scans > 0 else 0)
    
    return jsonify({
        'total_scans': total_scans,
        'ok_scans': ok_scans,
        'nok_scans': nok_scans,
        'success_rate': success_rate,
        'recent_scans': len(recent_scans)
    })


@app.route('/api/export/excel')
def export_excel():
    """API endpoint to export data to Excel"""
    status_filter = request.args.get('status', 'OK')
    search_term = request.args.get('search')
    log_file = request.args.get('log_file')
    fields_param = request.args.get('fields')

    data = log_parser.get_all_data(status_filter, search_term, log_file)

    column_definitions = [
        ('id_scan', 'ID Scan'),
        ('container_no', 'Nomor Container'),
        ('scan_time', 'Jam Scan'),
        ('update_time', 'Jam Update'),
        ('time_difference', 'Selisih Waktu'),
        ('image_count', 'Jumlah Gambar'),
        ('status', 'Status'),
        ('error_description', 'Deskripsi Error')
    ]
    column_map = {key: label for key, label in column_definitions}
    default_fields = [key for key, _ in column_definitions if key != 'error_description']

    requested_fields = []
    if fields_param:
        for field in fields_param.split(','):
            field_key = field.strip()
            if field_key and field_key in column_map and field_key not in requested_fields:
                requested_fields.append(field_key)

    if not requested_fields:
        requested_fields = default_fields

    headers = [column_map[field] for field in requested_fields]

    wb = Workbook()
    ws = wb.active
    if status_filter:
        sheet_title = f"Transmission {status_filter.upper()} Data"
    else:
        sheet_title = 'Transmission Log Data'
    ws.title = sheet_title[:31]

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    for col_index, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row_index, entry in enumerate(data, 2):
        for col_index, field in enumerate(requested_fields, 1):
            value = entry.get(field, '')
            if value is None:
                value = ''
            ws.cell(row=row_index, column=col_index, value=value)

    for col_index in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_index)
        max_length = 0
        for row_index in range(1, len(data) + 2):
            cell_value = ws[f"{column_letter}{row_index}"].value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename_status = (status_filter or 'all').lower()
    filename = f"transmission_log_{filename_status}_{timestamp}.xlsx"

    return send_file(
        temp_file.name,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/api/settings')
def get_settings():
    """API endpoint to get current settings"""
    return jsonify(app_settings)


@app.route('/api/ftp-status')
def get_ftp_status():
    """API endpoint to get cached FTP statuses."""
    with ftp_status_lock:
        status_snapshot = copy.deepcopy(ftp_status_cache)

    return jsonify({
        'statuses': status_snapshot,
        'ping_interval': app_settings.get('ftp_ping_interval',
                                          DEFAULT_FTP_PING_INTERVAL)
    })


@app.route('/api/ftp-status/ping', methods=['POST'])
def ping_ftp_status():
    """API endpoint to force an immediate FTP status check."""
    try:
        statuses = ftp_monitor.poll_now()
    except Exception as exc:  # pragma: no cover - defensive logging
        logger.exception("Failed to poll FTP status on demand: %s", exc)
        return jsonify({'error': 'Failed to poll FTP status'}), 500

    return jsonify({
        'statuses': statuses,
        'ping_interval': app_settings.get('ftp_ping_interval',
                                          DEFAULT_FTP_PING_INTERVAL)
    })


@app.route('/api/settings', methods=['POST'])
def update_settings():
    """API endpoint to update settings"""
    global app_settings, log_parser, ftp_monitor

    try:
        new_settings = request.get_json(silent=True)
        if not isinstance(new_settings, dict):
            return jsonify({'error': 'Invalid JSON payload'}), 400

        sanitized_settings = {}
        settings_changed = False

        if 'logs_directory' in new_settings:
            logs_dir = new_settings['logs_directory']
            if not logs_dir or not os.path.exists(logs_dir):
                return jsonify({'error': 'Logs directory does not exist'}), 400
            sanitized_settings['logs_directory'] = logs_dir

        if 'auto_refresh_interval' in new_settings:
            try:
                auto_refresh = int(new_settings['auto_refresh_interval'])
            except (TypeError, ValueError):
                return jsonify({
                    'error': 'Auto refresh interval must be a positive integer'
                }), 400

            if auto_refresh <= 0:
                return jsonify({
                    'error': 'Auto refresh interval must be a positive integer'
                }), 400

            sanitized_settings['auto_refresh_interval'] = auto_refresh

        if 'ftp_targets' in new_settings:
            try:
                ftp_targets = sanitize_ftp_targets(
                    new_settings['ftp_targets'], strict=True)
            except ValueError as exc:
                return jsonify({'error': str(exc)}), 400

            sanitized_settings['ftp_targets'] = ftp_targets

        if 'ftp_ping_interval' in new_settings:
            try:
                ftp_interval = validate_ping_interval(
                    new_settings['ftp_ping_interval'])
            except ValueError as exc:
                return jsonify({'error': str(exc)}), 400

            sanitized_settings['ftp_ping_interval'] = ftp_interval

        if 'resend_server' in new_settings:
            sanitized_settings['resend_server'] = str(
                new_settings['resend_server'] or ''
            ).strip()

        if 'resend_endpoint' in new_settings:
            sanitized_settings['resend_endpoint'] = str(
                new_settings['resend_endpoint'] or ''
            ).strip()

        if not sanitized_settings:
            return jsonify({'message': 'No settings were changed'}), 200

        if 'logs_directory' in sanitized_settings:
            logs_dir = sanitized_settings['logs_directory']
            if app_settings.get('logs_directory') != logs_dir:
                log_parser = LogParser(logs_dir)
                configure_ping_logger(logs_dir)
                settings_changed = True

        if 'auto_refresh_interval' in sanitized_settings:
            if app_settings.get('auto_refresh_interval') != sanitized_settings['auto_refresh_interval']:
                settings_changed = True

        if 'ftp_targets' in sanitized_settings:
            if app_settings.get('ftp_targets') != sanitized_settings['ftp_targets']:
                settings_changed = True

        if 'ftp_ping_interval' in sanitized_settings:
            if app_settings.get('ftp_ping_interval') != sanitized_settings['ftp_ping_interval']:
                settings_changed = True

        if 'resend_server' in sanitized_settings:
            if app_settings.get('resend_server', '') != sanitized_settings['resend_server']:
                settings_changed = True

        if 'resend_endpoint' in sanitized_settings:
            if app_settings.get('resend_endpoint', '') != sanitized_settings['resend_endpoint']:
                settings_changed = True

        app_settings.update(sanitized_settings)

        if not save_settings(app_settings):
            return jsonify({'error': 'Failed to save settings'}), 500

        if settings_changed:
            ftp_monitor.start(app_settings)

        return jsonify({'message': 'Settings updated successfully'})

    except Exception as e:  # pragma: no cover - defensive
        logger.exception("Failed to update settings: %s", e)
        return jsonify({'error': str(e)}), 400


@app.route('/api/validate-directory')
def validate_directory():
    """API endpoint to validate if a directory exists and contains log files"""
    directory = request.args.get('directory', '')
    
    if not directory:
        return jsonify({'valid': False, 'message': 'Directory path is required'})
    
    if not os.path.exists(directory):
        return jsonify({'valid': False, 'message': 'Directory does not exist'})
    
    if not os.path.isdir(directory):
        return jsonify({'valid': False, 'message': 'Path is not a directory'})
    
    # Check for log files
    pattern = os.path.join(directory, "Transmission.log*")
    log_files = glob.glob(pattern)
    
    if not log_files:
        return jsonify({
            'valid': False, 
            'message': 'No Transmission log files found in this directory'
        })
    
    return jsonify({
        'valid': True, 
        'message': f'Found {len(log_files)} log file(s)',
        'log_files': [os.path.basename(f) for f in log_files]
    })


def _extract_shutdown_token():
    header_token = request.headers.get('X-Controller-Token')
    if header_token:
        return header_token

    json_payload = request.get_json(silent=True)
    if isinstance(json_payload, dict) and 'token' in json_payload:
        return json_payload['token']

    if 'token' in request.args:
        return request.args['token']

    return None


@app.route('/__controller__/shutdown', methods=['POST'])
def controller_shutdown():
    """Internal endpoint invoked by the desktop controller to stop the server."""
    expected_token = app.config.get('SHUTDOWN_TOKEN', SHUTDOWN_TOKEN)
    provided_token = _extract_shutdown_token()

    if expected_token and provided_token != expected_token:
        return jsonify({'status': 'unauthorized'}), 403

    shutdown_event = app.config.get('SHUTDOWN_EVENT')
    if shutdown_event:
        shutdown_event.set()
        return jsonify({'status': 'stopping'})

    return jsonify({'status': 'no-op'}), 202


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
